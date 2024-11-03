import openpyxl
# path='C://Users//shibi//PycharmProjects//pythonProject3//TestData//LoginData.xlsx'
def getRowCount(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return (sheet.max_row)


def getColummnCount(file,sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    return (sheet.max_column)

def readData(file,sheetName,rownum,columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row=rownum,column=columnno).value

def writeData(file,sheetName,columnno,data,rownum):
    workbook = openpyxl.load_workbook(file)
    sheet=workbook[sheetName]
    sheet.cell(row=rownum,column=columnno).value=data
    workbook.save(file)

# rows=getRowCount(path,"Sheet1")
# for r in range(1, rows + 1):
#     user = readData(path, 'Sheet1', r, 1)
#     passw = readData(path, 'Sheet1', r, 2)
#     esp = readData(path, 'Sheet1', r, 3)
#
#     print(user)
#     print(passw)
#     print(esp)
#
# print(getColummnCount(path,"Sheet1"))